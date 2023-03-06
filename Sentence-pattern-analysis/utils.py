import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def read_xls(inxlsFile):
    extension = inxlsFile[-4:].upper()

    if extension == 'XLSX':
        wb = load_workbook(inxlsFile)  # ????????????wb not used?
        # Work sheet
        # data frame
        # df = pd.read_excel(inxlsFile, engine='openpyxl', sheet_name=selected_ws, header=start_r-1)

        f = pd.ExcelFile(inxlsFile).sheet_names
        # 'SUPPQUAL' not in  'DSTERM SUPPQUAL CONTENT Revision History'

        ff = [f[i] for i in range(len(f)) \
              if ('TEST' not in f[i]) and ('_' not in f[i]) \
              and (f[i] not in 'DSTERM SUPPQUAL CONTENT Revision History')
              ]

        # selected_ws = expander_ext.selectbox('Select Domain', ff)
        # selected_ws = 'AE'
        # targetDomain = selected_ws

        # Start from row
        # max_row = wb.worksheets[wb.sheetnames.index(selected_ws)].max_row
        # start_r = expander_ext.number_input('Start from row', 1, max_row)

        # f = pd.ExcelFile(inxlsFile).sheet_names
        f2 = [(i, ff[i]) for i in range(len(ff))]

        df = pd.read_excel(inxlsFile, sheet_name=ff, header=12)

        df0 = pd.read_excel(inxlsFile, sheet_name=ff, header=None, nrows=11)

        def trans(df0, domain):
            test = df0.get(domain)
            t = test.T
            t.iloc[3, 0] = t.iloc[0, 0].split(' ')[0]
            t.iloc[0, 0] = 'Domain Mapping Specifications'

            t = t.iloc[[0, 3],]
            t.columns = t.iloc[0]  # first row as column name
            t = t.iloc[[1], 1:]
            t["Domain"] = domain
            return t

        for f in ff:
            df.get(f)["Domain"] = f

        df_all = pd.concat([df.get(f) for f in ff])
        df0_all = pd.concat([trans(df0, f) for f in ff])

        df0_all = df0_all.reset_index()

    return df_all, df0_all


class GetVar():
    def __init__(self, s):
        self.raw_str = s
        self.operation = "<=|>=|=|\+|\-|\*\*|/|\*"
        # 预处理：计算符号中间空格去除，两边加空格
        self.s = self.pre_process(self.raw_str)
        self.s = self.replace_operations(self.s)
    def pre_process(self, s):
        # 去除逻辑符号之间多余的空格
        s = re.sub('([\*<=+-])( +)([\*<=+-])', lambda x: x.group(1)+x.group(3), s)
        s = re.sub('(<=|>=|<|>|=|\+|\-|\*\*|/|\*)', lambda x: ' ' + x.group(1) + ' ', s)
        s = s.strip(' ')

        if s[-1] == '.':
            s = s[:-1]+' '+'.'

        s = ' '+s+' '
        s = re.sub('[ ]+', ' ', s)

        for i, j in enumerate(s):
            if i == 0 or i==len(s)-1:
                continue
            if j == '.' and re.match('[A-Z]{1}',s[i-1]) and re.match('[^A-Z]{1}',s[i+1]):
                s = s[:i]+' '+s[i:]

        s = re.sub('\(', '( ', s)
        s = re.sub('\)', ' )', s)
        return s
    def replace_operations(self, s):
        return re.sub(r"({})+".format(self.operation), lambda x:' '+x.group(1)+' ', s)

    def var_rep(self, s=None):
        # 是变量的情况：1.连续大写字符 2.连续大写字符中间带点号 3.note中不连续大写字符（中间有空格）4.=号后面不连续大写字符（中间有空格）
        # 对于连续带空格大写字符，可以替换非单词字符后，判断是不是一句正常的语句
        if s is None:
            s = self.s

        # 某些特殊全部大写的字符，并不是变量
        extra_up = [' EPOCH ', ' NULL ']
        extra_low = [str.lower(x) for x in extra_up]
        for i in range(len(extra_up)):
            s = s.replace(extra_up[i], extra_low[i])

        # vote = re.findall('[\'\"].*[\'\"]', s)
        # pro_str = re.sub('[\'\"].*[\'\"]', '[note]')
        res = re.finditer('[^a-z]([A-Z]+[A-Z0-9._ ]+)[^a-z]', s)
        inds = []
        list_var = []
        for i in res:
            span = list(i.span())
            span[0] += 1
            span[1] -= 1
            inds.append(span)
        for x in inds:
            if x[1]<len(s)-1 and s[x[1]]=='(':
                if '.' in s[x[0]:x[1]]:
                    list_var.append(s[x[0]:x[1]])
                else:
                    pass
            else:
                list_var.append(s[x[0]:x[1]])

        list_var = [x.strip(' ') for x in list_var]
        # 需要长的字符先替换
        len_var = [len(i) for i in list_var]
        sort_ind = np.argsort(len_var)[::-1]
        for i in sort_ind:
            s=s.replace(list_var[i], ' [V] ')
        pro_str = s.strip(' ')
        pro_str = re.sub('  ', ' ', pro_str)
        n = len(list_var)
        var_dict = {}
        for i in range(1, n+1):
            var_dict['[V{}]'.format(i)] = list_var[i-1]
        k = 1
        pro_str_rep = ''
        i = 0
        last = 0
        while i<len(pro_str)-3:
            if pro_str[i:i+3] == '[V]':
                pro_str_rep += '[V{}]'.format(k)
                k+=1
                i+=3
                last = i - 1
            else:
                pro_str_rep += pro_str[i]
                i+=1
        if pro_str[-3:] == '[V]':
            pro_str_rep += '[V{}]'.format(k)
        else:
            if last >= len(pro_str) - 3:
                pro_str_rep += pro_str[last+1:]
            else:
                pro_str_rep += pro_str[-3:]

        # 变量以点号结尾的情况
        for i in var_dict:
            if var_dict[i][-1] == '.':
                var_dict[i] = var_dict[i][:-1]
                pro_str_rep = pro_str_rep.replace(i, i+'.')
            var_dict[i] = var_dict[i].strip(' ')
        return var_dict, pro_str_rep


# 处理引号
def sample_preprocess(s):
    # 对训练和预测样本进行预处理
    # 分号后面加空格
    if s == '':
        return s
    s = re.sub(';', '; ', s)
    s = s.replace('||', ' || ')
    s = s.replace('(', ' (')
    s = s.replace(')', ') ')
    s = s.replace("”", '"')
    re.sub('([A-Z]{1})\.[ ]{1}', lambda x: x.group(1) + ' . ', s)
    # 多个空格变成一个
    s = re.sub('[ ]+', ' ', s)
    return s


def sample_preprocess(s):
    # 对训练和预测样本进行预处理
    # 分号后面加空格
    if s == '':
        return s
    s = re.sub(';', '; ', s)
    s = s.replace('||', ' || ')
    s = s.replace('(', ' (')
    s = s.replace(')', ') ')
    # 多个空格变成一个
    s = re.sub('[ ]+', ' ', s)
    return s


def get_ents_rep(raws_list, ner_predict_list):
    '''
    实体用特殊字符替换
    :param raws_list:
    :param ner_predict_list: example:[('a+b',3,6,'FORMULA')]
    :return:
    '''
    reps_dict = dict()
    ents_dict = {'FUNC': 0, 'FORMULA': 0, 'FORMAT': 0, 'VAR': 0, 'QUOT': 0}
    ss = []
    for raw, entities in zip(raws_list, ner_predict_list):
        try:
            entities = eval(entities)
        except:
            pass
        inds, cats, ents= [0], [], []
        for ent in entities:
            ents.append(ent[0])
            inds.extend([ent[1], ent[2]])
            cats.append(ent[3])
        inds.append(-1)
        # ent_map = {'FUNC': '<function>', 'FORMULA': '<formula>', 'FORMAT': '<format>', 'VAR': '<variable>',
        #            'QUOT': '<quotation>'}
        # cats = [ent_map[x] for x in cats]
        m = len(cats)
        k = 0
        s = ''
        for i in range(0, len(inds), 2):
            if k <= m - 1:
                key = cats[k]+str(ents_dict[cats[k]])
                reps_dict[key] = ents[k]
                s += raw[inds[i]:inds[i + 1]] + key
                ents_dict[cats[k]] += 1
            k += 1
        try:
            s += raw[inds[-2]:]
        except:
            print(raw, inds)
            return
        ss.append(s)
    return ss, reps_dict


# 处理引号
def process_quotation(s):
    abbre = "'m |'s |s' |'t "
    if re.findall("\"{1}\'\"{1}", s) or re.findall("\"{1}\'\"{1}", s):
        print("quotation problem")
        return {},s

    # 获取abbre字符所在位置和值
    res = re.finditer(abbre, s)
    inds = []
    list_var = []
    for i in res:
        span = list(i.span())
        span[1] -= 1
        if s[span[0]] == "'":
            inds.append(span[0])
        else:
            inds.append(span[1]-1)
        list_var.append(s[span[0]:span[1]])


    # s'即可能是表示从属，也可能是引用符号，当表示引用符号时，从记录中删除
    del_ls = []
    for i, x in enumerate(list_var):
        if x=="s'":
            ind = inds[i]
            k = ind
            while k>0 and s[k]!=' ':
                k -= 1
                if s[k] == "'":
                   del_ls.append(i)
    inds = np.delete(inds, del_ls).tolist()

    # 记录引号位置
    quots = {}
    inds = []
    for i, v in enumerate(s):
        if v=='\'' or v== '"':
            if i not in inds:
                inds.append(i)

    n = len(inds)
    if n % 2 == 1:
        print("quotation problem when process {}".format(s))
        return {}, s
    if n == 0:
        return {}, s
    rep_s = ''
    for i in range(n//2):
        quots['[quot{}]'.format(i+1)] = s[inds[2*i]:inds[2*i+1]+1]
        if i == 0:
            rep_s = rep_s + s[:inds[2*i]] + ' [quot{}] '.format(i+1)
        else:
            rep_s = rep_s + s[inds[2*i-1]+1:inds[2 * i]] + ' [quot{}] '.format(i+1)
    if inds[2*i+1]<len(s)-1:
        rep_s += s[inds[2*i+1]+1:]
    return quots, rep_s