import numpy as np
import pandas as pd
# from nltk.corpus import wordnet
# import nltk
import re
# nltk.download('wordnet')
from openpyxl import load_workbook


def read_xls(inxlsFile):
    extension = inxlsFile[-4:].upper()

    if extension == 'XLSX':
        wb = load_workbook(inxlsFile)  # ????????????wb not used?
        # Work sheet
        # data frame
        # df = pd.read_excel(inxlsFile, engine='openpyxl', sheet_name=selected_ws, header=start_r-1)

        f = pd.ExcelFile(inxlsFile).sheet_names
        # 'SUPPQUAL' not in  'DSTERM SUPPQUAL CONTENT Revision History'

        ff = [f[i] for i in range(len(f)) \
              if ('TEST' not in f[i]) and ('_' not in f[i]) \
              and (f[i] not in 'DSTERM SUPPQUAL CONTENT Revision History')
              ]

        # selected_ws = expander_ext.selectbox('Select Domain', ff)
        # selected_ws = 'AE'
        # targetDomain = selected_ws

        # Start from row
        # max_row = wb.worksheets[wb.sheetnames.index(selected_ws)].max_row
        # start_r = expander_ext.number_input('Start from row', 1, max_row)

        # f = pd.ExcelFile(inxlsFile).sheet_names
        f2 = [(i, ff[i]) for i in range(len(ff))]

        df = pd.read_excel(inxlsFile, sheet_name=ff, header=12)

        df0 = pd.read_excel(inxlsFile, sheet_name=ff, header=None, nrows=11)

        def trans(df0, domain):
            test = df0.get(domain)
            t = test.T
            t.iloc[3, 0] = t.iloc[0, 0].split(' ')[0]
            t.iloc[0, 0] = 'Domain Mapping Specifications'

            t = t.iloc[[0, 3],]
            t.columns = t.iloc[0]  # first row as column name
            t = t.iloc[[1], 1:]
            t["Domain"] = domain
            return t

        for f in ff:
            df.get(f)["Domain"] = f

        df_all = pd.concat([df.get(f) for f in ff])
        df0_all = pd.concat([trans(df0, f) for f in ff])

        df0_all = df0_all.reset_index()

    return df_all, df0_all


class GetFormulaFunc():
    def __init__(self, s):
        self.raw_str = self.pre_process(s)
        self.punctuation = '!#,:;?@\\`~'
        self.operation = "<=|>=|=|\+|\-|\*\*|/|\*"
        self.operation_rep = '[op]'
        self.punc_rep = '[punc]'
        self.default_funcs = ['substr', 'vvalue', 'missing', 'in', 'upcase']
    def pre_process(self, s):
        # 去除逻辑符号之间多余的空格
        s = re.sub('([\*<=+-])( +)([\*<=+-])', lambda x: x.group(1)+x.group(3), s)
        return s

    def find_func(self, s=None):
        # 检测函数（函数名后面带括号）
        if s==None:
            s = self.raw_str
        l_ind = []
        r_ind = []
        left = re.finditer('\(', s)
        right = re.finditer('\)', s)
        for i in left:
            l_ind.append(i.span()[0])
        for i in right:
            r_ind.append(i.span()[0])

        n = len(l_ind)
        # 这里没有检查括号的合法性，只检查了数量对等性
        if n != len(r_ind):
            raise RuntimeError("bracket exception")
        if n == 0:
            return [],{}
        pair = {}
        i, j = 0, 0

        # 寻找括号对
        temp = []
        while True:
            if l_ind[i] < r_ind[j]:
                temp.append(l_ind[i])
                i += 1
                if i == n:
                    for val in r_ind[j:n]:
                        key = temp.pop(-1)
                        pair[key] = val
                    break
            else:
                key = temp.pop(-1)
                pair[key] = r_ind[j]
                j += 1

        # 计算每个函数的下标起始和结束位置
        funcs_ind = {}
        for i in pair:
            k = i
            bracket_ind = [k, pair[k]]
            if k == 0:
                continue

            # 如果括号前面为空格，判断前面的字符是否在默认函数列表
            if s[k-1] == ' ':
                k -= 2
                while k > 0  and s[k] != ' ':
                    k -= 1
                if k == 0:
                    k -= 1
                if s[k+1: i-1] in self.default_funcs:
                    funcs_ind[k+1] = bracket_ind
                continue
            if bool(re.match("[A-Za-z0-9_]", s[k-1])) == False:
                continue
            k -= 1
            while k>0 and bool(re.match("[A-Za-z0-9_]", s[k-1])):
                k -= 1
            if k - 2 > 0  and bool(re.match("[A-Z]{1}\.", s[k-2:k])):
                continue
            if bool(re.match("^[0-9]", s[k: bracket_ind[0]])):
                continue

            funcs_ind[k] = bracket_ind

        # 返回替换函数后的字符串和函数字符串
        func_names = []
        for i, start_ind in enumerate(funcs_ind):
            func_names.append(s[start_ind: funcs_ind[start_ind][0]])
        return func_names, funcs_ind

    # punctuation
    def remove_punctuation(self, s):
        return re.sub(r"[{}]+".format(self.punctuation), " ", s)

    def findall_punctuation(self, s):
        return re.findall(r"[{}]+".format(self.punctuation), s)

    def replace_punctuations(self, s):
        return re.sub(r"[{}]+".format(self.punctuation), ' {} '.format(self.punc_rep), s)

    # operations
    def replace_operations(self, s):
        return re.sub(r"({})+".format(self.operation), ' {} '.format(self.operation_rep), s)

    def findall_operation(self, s):
        return re.findall(r"({})+".format(self.operation), s)

    def findall_formula(self, s=None, number=0):
        if s is None:
            s = self.raw_str

        # 去除多余标点
        list_punc = self.findall_punctuation(s)
        pro_str = self.replace_punctuations(s)

        # 找到并替换运算符号
        list_op = self.findall_operation(pro_str)
        pro_str = self.replace_operations(pro_str)

        len_op = len(list_op)
        str_split = pro_str.split()

        # 获取公式字符及其下标值
        formulas = []
        inds = []
        ind = 0
        k_op = 0
        n = len(str_split)
        while ind < n - 1 and k_op < len_op:
            formula = []
            while str_split[ind] == self.operation_rep:
                if ind == 0:
                    formula.append(list_op[k_op])
                    inds.append(ind)
                else:
                    formula.append(str_split[ind - 1])
                    formula.append(list_op[k_op])
                    inds.extend([ind-1, ind])
                le = len(re.findall('\(', str_split[ind+1]))
                if le == 0:
                    pass
                else:
                    while True:
                        ri = len(re.findall('\)', str_split[ind+1]))
                        diff = le - ri
                        if diff <= 0 or ind + 1>n-1:
                            break
                        if str_split[ind+1] == self.operation_rep:
                            formula.append(list_op[k_op])
                            k_op += 1
                        else:
                            formula.append(str_split[ind+1])
                        inds.append(ind+1)
                        ind += 1
                        le += len(re.findall('\(', str_split[ind + 1]))
                        ri += len(re.findall('\)', str_split[ind + 1]))
                k_op += 1
                ind = ind + 2
                if ind >= n - 1:
                    break
            if len(formula) > 0:
                formula.append(str_split[ind - 1])
                inds.append(ind-1)
                formulas.append(formula)
            ind += 1
        formulas = [' '.join(x) for x in formulas]

        i = 0
        formula_ind_dict = {} # 只记录公式出现的开始结束位置
        while i < len(inds)-1:
            key = inds[i]
            while i<len(inds)-1 and inds[i+1] - inds[i] == 1:
                i += 1
            formula_ind_dict[key] = inds[i]
            i += 1
        formula_num = len(formula_ind_dict)

        # 公式用'[formu1],[formu2]...'代替
        sorted_keys = sorted(formula_ind_dict)
        res = []
        if formula_num == 0:
            pass
        elif formula_num == 1:
            res = str_split[:sorted_keys[0]]+['[formu{}]'.format(number+1)]+str_split[formula_ind_dict[sorted_keys[0]]+1:]
        elif formula_num == 2:
            res.extend(str_split[:sorted_keys[0]]+['[formu{}]'.format(number+1)])
            res.extend(str_split[formula_ind_dict[sorted_keys[0]]+1:sorted_keys[1]]+['[formu{}]'.format(number+2)])
            res.extend(str_split[formula_ind_dict[sorted_keys[1]]+1:])
        else:
            for i, key in enumerate(sorted_keys):
                if i == 0:
                    res.extend(str_split[:key])
                    res.append('[formu{}]'.format(number+i+1))
                elif i == formula_num - 1:
                    res.extend(str_split[formula_ind_dict[sorted_keys[i-1]]+1:key])
                    res.append('[formu{}]'.format(number + i + 1))
                    if formula_ind_dict[key]+1 < n:
                        res.extend(str_split[formula_ind_dict[key]+1:])
                else:
                    res.extend(str_split[formula_ind_dict[sorted_keys[i-1]]+1:key])
                    res.append('[formu{}]'.format(number + i + 1))

        # 标点符号还原
        for i, v in enumerate(res):
            if v == self.punc_rep:
                punc = list_punc.pop(0)
                res[i] = punc

        res = ' '.join(res)

        # 为公式建立索引
        formulas_dict = {}

        for i, key in enumerate(sorted_keys):
            formulas_dict['[formu{}]'.format(number+i+1)]=formulas[i]

        return formulas_dict, res

    def recur_rep(self, s=None):
        # 从内到外，逐步检测出公式、函数
        if s == None:
            s = self.raw_str
        # 找到函数
        _, func_ind = self.find_func(self.raw_str)
        '''
        # 为了防止括号旁边的公式内容检测出问题，在括号旁边加空格
        s_list = list(self.raw_str)
        for i in func_ind:
            s_list[func_ind[i][0]] = '( '
            s_list[func_ind[i][1]] = ' )'
        pro_str = ''.join(s_list)
        '''
        sorted_keys = sorted(func_ind, reverse=True)

        number = 0
        all_formulas_dict = {}
        all_funcs_dict = {}
        if len(sorted_keys) == 0:
            pass
        else:
            for i, key in enumerate(sorted_keys):
                formulas_dict, res_str = self.findall_formula(s[func_ind[key][0]+1: func_ind[key][1]], number=number)
                if len(formulas_dict) == 0:
                    temp = s[key: func_ind[key][1]+1]
                    temp = re.sub('~*','',temp)
                    all_funcs_dict['[fc{}]'.format(i)] = temp
                    l_diff = func_ind[key][1] - key + 1 - len('[fc{}]'.format(i))
                    s = s[:key] + '[fc{}]'.format(i) + '~' * l_diff + s[func_ind[key][1] + 1:]
                else:
                    number += len(formulas_dict)
                    all_formulas_dict.update(formulas_dict)

                    # 替换后产生的字符长度差
                    l_diff = func_ind[key][1] - key + 1 - len('[fc{}]'.format(i))

                    temp = s[key: func_ind[key][0]+1]+res_str+')'
                    temp = re.sub('~*', '', temp)
                    all_funcs_dict['[fc{}]'.format(i)] = temp
                    s = s[:key]+'[fc{}]'.format(i)+'~'*l_diff+s[func_ind[key][1]+1:]

        s = re.sub("~*",'', s)
        formulas_dict, res_str = self.findall_formula(s, number=number)
        if len(formulas_dict) == 0:
            return all_formulas_dict, all_funcs_dict, s
        else:
            all_formulas_dict.update(formulas_dict)
            return all_formulas_dict, all_funcs_dict, res_str


class GetVar():
    def __init__(self, s):
        self.raw_str = s
        self.operation = "<=|>=|=|\+|\-|\*\*|/|\*"
        # 预处理：计算符号中间空格去除，两边加空格
        self.s = self.pre_process(self.raw_str)
        self.s = self.replace_operations(self.s)
    def pre_process(self, s):
        # 去除逻辑符号之间多余的空格
        s = re.sub('([\*<=+-])( +)([\*<=+-])', lambda x: x.group(1)+x.group(3), s)
        s = s.strip(' ')

        if s[-1] == '.':
            s = s[:-1]+' '+'.'

        s = ' '+s+' '

        for i, j in enumerate(s):
            if i == 0 or i==len(s)-1:
                continue
            if j == '.' and re.match('[A-Z]{1}',s[i-1]) and re.match('[^A-Z]{1}',s[i+1]):
                s = s[:i]+' '+s[i:]

        s = re.sub('\(', '( ', s)
        s = re.sub('\)', ' )', s)
        return s
    def replace_operations(self, s):
        return re.sub(r"({})+".format(self.operation), lambda x:' '+x.group(1)+' ', s)

    def var_rep(self, s=None):
        # 是变量的情况：1.连续大写字符 2.连续大写字符中间带点号 3.note中不连续大写字符（中间有空格）4.=号后面不连续大写字符（中间有空格）
        # 对于连续带空格大写字符，可以替换非单词字符后，判断是不是一句正常的语句
        if s is None:
            s = self.s

        # 某些特殊全部大写的字符，并不是变量
        extra_up = [' EPOCH ', ' NULL ']
        extra_low = [str.lower(x) for x in extra_up]
        for i in range(len(extra_up)):
            s = s.replace(extra_up[i], extra_low[i])

        # vote = re.findall('[\'\"].*[\'\"]', s)
        # pro_str = re.sub('[\'\"].*[\'\"]', '[note]')
        res = re.finditer('[^a-z]([A-Z]+[A-Z0-9._ ]+)[^a-z]', s)
        inds = []
        list_var = []
        for i in res:
            span = list(i.span())
            span[0] += 1
            span[1] -= 1
            inds.append(span)
        for x in inds:
            if x[1]<len(s)-1 and s[x[1]]=='(':
                if '.' in s[x[0]:x[1]]:
                    list_var.append(s[x[0]:x[1]])
                else:
                    pass
            else:
                list_var.append(s[x[0]:x[1]])

        list_var = [x.strip(' ') for x in list_var]
        # 需要长的字符先替换
        len_var = [len(i) for i in list_var]
        sort_ind = np.argsort(len_var)[::-1]
        for i in sort_ind:
            s=s.replace(list_var[i], ' [V] ')
        pro_str = s.strip(' ')
        pro_str = re.sub('  ', ' ', pro_str)
        n = len(list_var)
        var_dict = {}
        for i in range(1, n+1):
            var_dict['[V{}]'.format(i)] = list_var[i-1]
        k = 1
        pro_str_rep = ''
        i = 0
        last = 0
        while i<len(pro_str)-3:
            if pro_str[i:i+3] == '[V]':
                pro_str_rep += '[V{}]'.format(k)
                k+=1
                i+=3
                last = i - 1
            else:
                pro_str_rep += pro_str[i]
                i+=1
        if pro_str[-3:] == '[V]':
            pro_str_rep += '[V{}]'.format(k)
        else:
            if last >= len(pro_str) - 3:
                pro_str_rep += pro_str[last+1:]
            else:
                pro_str_rep += pro_str[-3:]

        # 变量以点号结尾的情况
        for i in var_dict:
            if var_dict[i][-1] == '.':
                var_dict[i] = var_dict[i][:-1]
                pro_str_rep = pro_str_rep.replace(i, i+'.')
        return var_dict, pro_str_rep


# 还需要此类引号：I'm/don't等等
# 处理引号
def process_quotation(s):
    abbre = "'m |'s |s' |'t "
    if re.findall("\"{1}\'\"{1}", s) or re.findall("\"{1}\'\"{1}", s):
        print("quotation problem")
        return {},s

    # 获取abbre字符所在位置和值
    res = re.finditer(abbre, s)
    inds = []
    list_var = []
    for i in res:
        span = list(i.span())
        span[1] -= 1
        if s[span[0]] == "'":
            inds.append(span[0])
        else:
            inds.append(span[1]-1)
        list_var.append(s[span[0]:span[1]])


    # s'即可能是表示从属，也可能是引用符号，当表示引用符号时，从记录中删除
    del_ls = []
    for i, x in enumerate(list_var):
        if x=="s'":
            ind = inds[i]
            k = ind
            while k>0 and s[k]!=' ':
                k -= 1
                if s[k] == "'":
                   del_ls.append(i)
    inds = np.delete(inds, del_ls).tolist()

    # 记录引号位置
    quots = {}
    inds = []
    for i, v in enumerate(s):
        if v=='\'' or v== '"':
            if i not in inds:
                inds.append(i)

    n = len(inds)
    if n % 2 == 1:
        print("quotation problem when process {}".format(s))
        return {}, s
    if n == 0:
        return {}, s
    rep_s = ''
    for i in range(n//2):
        quots['[quot{}]'.format(i+1)] = s[inds[2*i]:inds[2*i+1]+1]
        if i == 0:
            rep_s = rep_s + s[:inds[2*i]] + ' [quot{}] '.format(i+1)
        else:
            rep_s = rep_s + s[inds[2*i-1]+1:inds[2 * i]] + ' [quot{}] '.format(i+1)
    if inds[2*i+1]<len(s)-1:
        rep_s += s[inds[2*i+1]+1:]
    return quots, rep_s


def process_sasformat(s):
    formats = re.findall('\$[A-Za-z]+\.', s)
    format_dict = {}
    for i, v in enumerate(formats):
        format_dict['[format{}]'.format(i+1)] = v
    n = len(formats)
    if n == 0:
        return {}, s
    res = re.finditer('\$[A-Za-z]+\.', s)
    inds = []
    rep_s = ''
    for i in res:
        inds.extend(i.span())
    for i in range(n):
        if i == 0:
            rep_s = rep_s + s[:inds[2 * i]] + ' [format{}] '.format(i+1)
        else:
            rep_s = rep_s + s[inds[2 * i - 1]: inds[2 * i]] + ' [format{}] '.format(i+1)
    if inds[2*i+1]<len(s)-1:
        rep_s += s[inds[2*i+1]:]
    return format_dict, rep_s


def preprocess(s):
    return re.sub('e\.g\.', 'for example', s)


# s1 = 'a + 1 - b, set it that. m*kda'
# s2 = 'm=var2,if a+1=2: TRCAT=upcase(put(CRITYPE+1,$ONCRCAT.))'
# s3="Set as 'FOLLOW-UP' if SE.SESTDTC< =AESTDTC <=SE.SEENDTC where SE.EPOCH= FOLLOW-UP;"
# s4="Calculates the day in relation to the variables (TR.TRDTC) and subject ref start date/time (DM.RFSTDTC).If date is prior to ref start date, then TR.TRDY = TR.TRDTC-DM.RFSTDTC. Otherwise TR.TRDY = TR.TRDTC - DM.RFSTDTC +1. "
# formu_dict,funcs_dict,res = GetFormulaFunc(s1).recur_rep()
# print(formu_dict,funcs_dict,res)


if __name__ == '__main__':
    # s = ' Set to " [V1] - " ||  [V2]  if  [V3]  is not missing . '
    # formu_dict, funcs_dict, res = GetFormulaFunc(s).recur_rep()

    # data, _ = read_xls('/Users/wkx/Desktop/AZ/formula_extract/D8850C00008_SDTM3.3_Mapping_Specifications_v1.0.xlsx')
    data, _ = read_xls('AZ/data/D8850C00008_SDTM3.3_Mapping_Specifications_v1.0.xlsx')

    data = data[['Domain', 'Variable Name', 'Conversion Definition']].dropna()
    # data = pd.read_excel('/Users/wkx/Desktop/AZ/formula_extract/data1.xlsx')

    # a, b=GetVar(s).var_rep()
    data['Conversion Definition'] = data['Conversion Definition'].apply(lambda x:preprocess(x))
    data = data[data['Conversion Definition']!=' ']

    # input string list
    ss = list(data['Conversion Definition'])

    # 引号引用内容查找替换
    quots, rep_quots = [], []
    for i in ss:
        quot, rep_quot = process_quotation(i)
        quots.append(quot)
        rep_quots.append(rep_quot)
    data['quotation'] = quots

    # sas format格式内容查找替换
    formats = []
    rep_formats = []
    for i in rep_quots:
        a, b = process_sasformat(i)
        formats.append(a)
        rep_formats.append(b)
    data['format'] = formats

    # 变量内容查找替换
    Vars = []
    reps_var = []
    for i in rep_formats:
        Var, rep = GetVar(i).var_rep()
        Vars.append(Var)
        reps_var.append(rep)
    data['Vars'] = Vars
    data['rep_var'] = reps_var

    # 公式和函数内容查找替换
    formu_dicts, funcs_dicts, reses = [], [], []
    for ind,i in enumerate(reps_var):
        try:
            formu_dict, funcs_dict, res = GetFormulaFunc(i).recur_rep()
            formu_dicts.append(formu_dict)
            funcs_dicts.append(funcs_dict)
            reses.append(res)
        except:
            formu_dicts.append({})
            funcs_dicts.append({})
            reses.append('')
    data['formu_dict']=formu_dicts
    data['funcs_dict'] = funcs_dicts
    data['res'] = reses

    # 分句
    data['sentence'] = data['res'].apply(lambda x: re.split(';|\.', x))
    data = data.explode('sentence')
    data = data[data['sentence'] != '']
    # data.to_excel('/Users/wkx/Desktop/AZ/formula_extract/res2.xlsx', index=False)
    data.to_excel('AZ/data/res2.xlsx', index=False)


