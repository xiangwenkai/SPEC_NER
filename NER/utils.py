import pandas as pd
from openpyxl import load_workbook

# import os
# import shutil
# import time
# import re
# from datetime import datetime

def read_xls(inxlsFile):
    extension = inxlsFile[-4:].upper()

    if extension == 'XLSX':
        wb = load_workbook(inxlsFile)  # ????????????wb not used?
        # Work sheet
        # data frame
        # df = pd.read_excel(inxlsFile, engine='openpyxl', sheet_name=selected_ws, header=start_r-1)

        f = pd.ExcelFile(inxlsFile).sheet_names
        # 'SUPPQUAL' not in  'DSTERM SUPPQUAL CONTENT Revision History'

        ff = [f[i] for i in range(len(f)) \
              if ('TEST' not in f[i]) and ('_' not in f[i]) \
              and (f[i] not in 'DSTERM SUPPQUAL CONTENT Revision History')
              ]

        # selected_ws = expander_ext.selectbox('Select Domain', ff)
        # selected_ws = 'AE'
        # targetDomain = selected_ws

        # Start from row
        # max_row = wb.worksheets[wb.sheetnames.index(selected_ws)].max_row
        # start_r = expander_ext.number_input('Start from row', 1, max_row)

        # f = pd.ExcelFile(inxlsFile).sheet_names
        f2 = [(i, ff[i]) for i in range(len(ff))]

        df = pd.read_excel(inxlsFile, sheet_name=ff, header=12)

        df0 = pd.read_excel(inxlsFile, sheet_name=ff, header=None, nrows=11)

        def trans(df0, domain):
            test = df0.get(domain)
            t = test.T
            t.iloc[3, 0] = t.iloc[0, 0].split(' ')[0]
            t.iloc[0, 0] = 'Domain Mapping Specifications'

            t = t.iloc[[0, 3],]
            t.columns = t.iloc[0]  # first row as column name
            t = t.iloc[[1], 1:]
            t["Domain"] = domain
            return t

        for f in ff:
            df.get(f)["Domain"] = f

        df_all = pd.concat([df.get(f) for f in ff])
        df0_all = pd.concat([trans(df0, f) for f in ff])

        df0_all = df0_all.reset_index()

    return df_all, df0_all